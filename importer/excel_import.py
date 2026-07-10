"""Einmal-/Re-Import der Bestaende aus Basisinfos/Assets.xlsx (siehe Spezifikation B-5)."""
from __future__ import annotations

import sqlite3
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

import config
import database.db as db

XLSX_PATH = Path(__file__).resolve().parent.parent / "Basisinfos" / "Assets.xlsx"
SHEET_NAME_KRYPTO = "Krypto"
# Non-Krypto-Sheet (Aktien/ETF/Rohstoffe, Multi-Asset-Tracking) - Design abgestimmt
# 2026-07-09, hier umgesetzt (2026-07-10): identische Spalten wie das Krypto-Sheet,
# gleiche read_holdings_from_excel()/export_holdings()-Logik. Optional in der
# Nutzer-Datei (P-10: fehlt das Sheet, wird es uebersprungen statt abzustuerzen -
# die Non-Krypto-Bestaende sind dann schlicht noch nicht erfasst).
SHEET_NAME_NICHT_KRYPTO = "Nicht-Krypto"
# Gegenstueck zum Import (Nutzer-Idee 2026-07-09): eine SEPARATE Datei, nie die
# handgepflegte Original-Assets.xlsx direkt ueberschreiben. Export erzeugt eine
# frische Arbeitsmappe (kein Risiko, bestehende Formatierung/Zusatzspalten zu
# zerstoeren, wie es ein In-Place-Rundlauf koennte) - der Nutzer kann sie pruefen/
# bearbeiten und via "Bestaende aus Datei importieren..." (Filedialog, ui/app.py)
# wieder einlesen, import_holdings() akzeptiert dafuer bereits einen path-Parameter.
EXPORT_XLSX_PATH = Path(__file__).resolve().parent.parent / "Basisinfos" / "Assets_export.xlsx"


@dataclass
class ImportResult:
    imported_count: int
    warnings: list[str] = field(default_factory=list)


def _parse_quantity(raw) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, str):
        raw = raw.strip().replace(",", ".")
        if not raw:
            return 0.0
        return float(raw)
    return float(raw)


def _read_sheet(sheet) -> dict[str, float]:
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    col_index = {name: idx for idx, name in enumerate(header)}
    coin_name_idx = col_index["Coin Name"]
    kurzzeichen_idx = col_index["Kurzzeichen"]
    anzahl_idx = col_index["Anzahl Coins"]

    holdings: dict[str, float] = {}
    for row in sheet.iter_rows(min_row=2):
        coin_name = row[coin_name_idx].value
        if coin_name is None:
            break
        raw_symbol = row[kurzzeichen_idx].value
        symbol = (raw_symbol or "").strip().upper()
        quantity = _parse_quantity(row[anzahl_idx].value)
        holdings[symbol] = quantity

    return holdings


def read_holdings_from_excel(path: Path = XLSX_PATH) -> dict[str, float]:
    """Liest Krypto- UND Nicht-Krypto-Sheet (falls vorhanden) und merged beide
    Bestandslisten. Das Nicht-Krypto-Sheet ist optional (P-10) - fehlt es in der
    Nutzer-Datei (noch nicht angelegt), wird es einfach uebersprungen, kein Absturz."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    holdings: dict[str, float] = {}
    holdings.update(_read_sheet(workbook[SHEET_NAME_KRYPTO]))
    if SHEET_NAME_NICHT_KRYPTO in workbook.sheetnames:
        holdings.update(_read_sheet(workbook[SHEET_NAME_NICHT_KRYPTO]))
    return holdings


def import_holdings(
    conn: sqlite3.Connection, path: Path = XLSX_PATH, source: str = "import"
) -> ImportResult:
    holdings = read_holdings_from_excel(path)
    watchlist_symbols = {asset.symbol for asset in config.get_watchlist()}
    # Konflikt-Check (Nutzer-Review 2026-07-09): holdings hat keine Historie, ein
    # Reimport ueberschreibt absolut. Wurde ein Bestand seither per Signal-
    # Rueckmeldung (ui/signals_view.py::UmsetzungDialog, source="signal_bestaetigung")
    # auf einen ANDEREN Wert gesetzt, geht dieser sonst kommentarlos verloren -
    # ueberschreibt weiterhin (Assets.xlsx bleibt die Quelle der Wahrheit), aber
    # sichtbar in der Warnliste statt still.
    existing_holdings = {h.symbol: h for h in db.get_all_holdings(conn)}

    warnings: list[str] = []
    for symbol, quantity in holdings.items():
        if symbol not in watchlist_symbols:
            warnings.append(
                f"Symbol '{symbol}' aus Assets.xlsx nicht in config.yaml watchlist gefunden."
            )
        existing = existing_holdings.get(symbol)
        if existing is not None and existing.source == "signal_bestaetigung" and existing.quantity != quantity:
            when = existing.updated_at[:16].replace("T", " ")
            warnings.append(
                f"Symbol '{symbol}': Bestand wurde zuletzt manuell per Signal-Rückmeldung auf "
                f"{existing.quantity} gesetzt (am {when}) und wird jetzt durch den Excel-Import auf "
                f"{quantity} überschrieben."
            )
        db.upsert_holding(conn, symbol, quantity, source=source)

    missing_in_xlsx = watchlist_symbols - set(holdings.keys())
    for symbol in sorted(missing_in_xlsx):
        warnings.append(f"Watchlist-Symbol '{symbol}' hat keinen Eintrag in Assets.xlsx.")

    db.mark_holdings_imported(conn)

    return ImportResult(imported_count=len(holdings), warnings=warnings)


def _write_sheet(sheet, assets, holdings_by_symbol) -> None:
    sheet.append(["Coin Name", "Kurzzeichen", "Anzahl Coins", "Quelle"])
    for asset in assets:
        holding = holdings_by_symbol.get(asset.symbol)
        quantity = holding.quantity if holding else 0.0
        quelle = holding.source if holding else "-"
        sheet.append([asset.name, asset.symbol, quantity, quelle])


def export_holdings(conn: sqlite3.Connection, path: Path = EXPORT_XLSX_PATH) -> int:
    """Schreibt den aktuellen holdings-Stand in eine neue Arbeitsmappe (gleiches
    Format wie read_holdings_from_excel() erwartet - zwei Sheets 'Krypto'/
    'Nicht-Krypto', Spalten 'Coin Name'/'Kurzzeichen'/'Anzahl Coins'/'Quelle'). Eine
    Zeile pro Watchlist-Asset (nicht nur pro vorhandenem Bestand), Bestand 0 fuer
    Assets ohne holdings-Eintrag - so ist die Datei beim naechsten Import
    vollstaendig und loest keine "kein Eintrag in Assets.xlsx"-Warnung aus. Gibt die
    Anzahl geschriebener Zeilen (beide Sheets zusammen) zurueck.

    Zwei Sheets nach Assetklasse getrennt (2026-07-10, Non-Krypto-Slice, Design
    bereits 2026-07-09 abgestimmt, siehe Basisinfos/Spezifikation.md) - Krypto- und
    Nicht-Krypto-Assets (Aktien/ETF/Rohstoffe) landen je in ihrem eigenen Sheet,
    identisches Spaltenformat wie zuvor bei "Krypto" allein.

    Spalte "Quelle" (2026-07-10, Bitpanda-Sync-Slice): reine Herkunfts-Anzeige
    (import/signal_bestaetigung/bitpanda_sync/-), wird beim Reimport ueber den
    Header-Namen-Lookup in read_holdings_from_excel() automatisch ignoriert - kein
    neuer Quelle-pro-Zeile-Importmechanismus, Rundlauf bleibt unveraendert."""
    holdings_by_symbol = {h.symbol: h for h in db.get_all_holdings(conn)}
    watchlist = sorted(config.get_watchlist(), key=lambda asset: asset.symbol)
    krypto_assets = [a for a in watchlist if a.assetklasse == "krypto"]
    nicht_krypto_assets = [a for a in watchlist if a.assetklasse != "krypto"]

    workbook = openpyxl.Workbook()
    krypto_sheet = workbook.active
    krypto_sheet.title = SHEET_NAME_KRYPTO
    _write_sheet(krypto_sheet, krypto_assets, holdings_by_symbol)

    nicht_krypto_sheet = workbook.create_sheet(SHEET_NAME_NICHT_KRYPTO)
    _write_sheet(nicht_krypto_sheet, nicht_krypto_assets, holdings_by_symbol)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return len(watchlist)
